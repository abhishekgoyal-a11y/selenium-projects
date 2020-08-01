from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
import openpyxl

driver = webdriver.Chrome(
    'C:/Users/HP/Desktop/Drivers/chromedriver_win32/chromedriver.exe')
driver.get('https://www.gsmarena.com/samsung-phones-9.php')
workbook = openpyxl.load_workbook('C:/Users/HP/Desktop/try1.xlsx')
sheet = workbook.active
row = sheet.max_row
col = sheet.max_column
driver.maximize_window()
numberofmobile = driver.find_elements_by_xpath(
    '//*[@id="review-body"]/div[1]/ul/li')
for i in range(1, len(numberofmobile)):
    try:
        mobilebutton = driver.find_element_by_xpath(
            f'//*[@id="review-body"]/div[1]/ul/li[{i}]/a/strong/span')
        mobilebutton.click()
        mobilename = driver.find_element_by_xpath(
            '//*[@id="body"]/div/div[1]/div/div[1]/h1')
        inch = driver.find_element_by_xpath(
            '//*[@id="body"]/div/div[1]/div/div[2]/ul/li[4]/strong/span')
        camera = driver.find_element_by_xpath(
            '//*[@id="body"]/div/div[1]/div/div[2]/ul/li[5]/strong/span[1]')
        ram = driver.find_element_by_xpath(
            '//*[@id="body"]/div/div[1]/div/div[2]/ul/li[6]/strong/span[1]')
        rom = driver.find_element_by_xpath(
            '//*[@id="body"]/div/div[1]/div/div[2]/ul/li[1]/span[4]/span')
        battery = driver.find_element_by_xpath(
            '//*[@id="body"]/div/div[1]/div/div[2]/ul/li[7]/strong/span[1]')
        sheet.cell(row=i, column=1).value = mobilename.text
        sheet.cell(row=i, column=2).value = inch.text
        sheet.cell(row=i, column=3).value = camera.text
        sheet.cell(row=i, column=4).value = ram.text
        sheet.cell(row=i, column=5).value = rom.text
        sheet.cell(row=i, column=6).value = battery.text
        print(mobilename.text, inch.text, camera.text,
              ram.text, rom.text, battery.text)
        driver.back()
    except:
        pass
    workbook.save('C:/Users/HP/Desktop/try1.xlsx')
