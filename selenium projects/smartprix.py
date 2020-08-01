from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
import openpyxl
workbook = openpyxl.load_workbook("C:/Users/HP/Desktop/try.xlsx")
sheet = workbook.active
row = sheet.max_row
col = sheet.max_column
driver = webdriver.Chrome(
    'C:/Users/HP/Desktop/Drivers/chromedriver_win32/chromedriver.exe')
driver.get('https://www.smartprix.com/mobiles/samsung-brand')
numberofmobile = driver.find_element_by_xpath(
    '//*[@id = "search-result-items"]/div[1]/div[1]/strong')
numberofmobileone = int(numberofmobile.text[0])
numberofmobiletwo = int(numberofmobile.text[2:4])
for i in range(numberofmobileone, numberofmobiletwo+1):
    name = driver.find_element_by_xpath(
        f'//*[@id="search-result-items"]/div[1]/ul/li[{i}]/div[3]/h2')
    ramrom = driver.find_element_by_xpath(
        f'//*[@id = "search-result-items"]/div[1]/ul/li[{i}]/div[3]/div/ul[1]/li[3]')
    battery = driver.find_element_by_xpath(
        f'//*[@id="search-result-items"]/div[1]/ul/li[{i}]/div[3]/div/ul[1]/li[4]')
    inch = driver.find_element_by_xpath(
        f'//*[@id="search-result-items"]/div[1]/ul/li[{i}]/div[3]/div/ul[2]/li[1]')
    camera = driver.find_element_by_xpath(
        f'//*[@id="search-result-items"]/div[1]/ul/li[{i}]/div[3]/div/ul[2]/li[2]')
    extendstorage = driver.find_element_by_xpath(
        f'//*[@id="search-result-items"]/div[1]/ul/li[{i}]/div[3]/div/ul[2]/li[3]')
    sheet.cell(row=i, column=1).value = name.text
    sheet.cell(row=i, column=2).value = ramrom.text[0:16]+" ROM"
    sheet.cell(row=i, column=3).value = battery.text[0:8]
    sheet.cell(row=i, column=4).value = inch.text[0:10]
    sheet.cell(row=i, column=5).value = camera.text
    sheet.cell(row=i, column=6).value = extendstorage.text[27:]
    print(name.text, ramrom.text, battery.text,
          inch.text[0:10], camera.text, extendstorage.text[27:])
workbook.save("C:/Users/HP/Desktop/try.xlsx")
driver.close()
