from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
import openpyxl
workbook = openpyxl.load_workbook("C:/Users/HP/Desktop/try.xlsx")
sheet = workbook.active
row = sheet.max_row
col = sheet.max_column
driver = webdriver.Chrome(
    'C:/Users/HP/Desktop/Drivers/chromedriver_win32/chromedriver.exe')
driver.get('https://www.flipkart.com/search?q=samsung+mobiles&sid=tyy%2C4io&as=on&as-show=on&otracker=AS_QueryStore_OrganicAutoSuggest_1_6_na_na_na&otracker1=AS_QueryStore_OrganicAutoSuggest_1_6_na_na_na&as-pos=1&as-type=RECENT&suggestionId=samsung+mobiles%7CMobiles&requestId=a7282b3a-c855-484f-8552-7fa705ad624e&as-searchtext=samsun')
driver.maximize_window()
numberofmobile = driver.find_element_by_xpath(
    '//*[@id="container"]/div/div[3]/div[2]/div[1]/div[2]/div[1]/div/div/span').text
numberofmobileone = int(numberofmobile[8:10])
numberofmobiletwo = int(numberofmobile[12:15])
numberofmobilethree = (numberofmobiletwo-(numberofmobileone-1))
numberofmobilefour = int(numberofmobile[18:21])
numberofmobilefive = numberofmobilefour//numberofmobiletwo
a = 2
b = 0
while a < numberofmobilefive:
    for i in range(2, numberofmobilethree+1):
        # mobilename = driver.find_element_by_xpath(
        #     f'//*[@id="container"]/div/div[3]/div[2]/div[1]/div[2]/div[{i}]/div/div/div/a/div[2]/div[1]/div[1]')
        # ramrom = driver.find_element_by_xpath(
        #     f'//*[@id="container"]/div/div[3]/div[2]/div[1]/div[2]/div[{i}]/div/div/div/a/div[2]/div[1]/div[3]/ul/li[1]')
        inch = driver.find_element_by_xpath(
            f'//*[@id="container"]/div/div[3]/div[2]/div[1]/div[2]/div[{i}]/div/div/div/a/div[2]/div[1]/div[3]/ul/li[2]')
        '//*[@id="container"]/div/div[3]/div[2]/div[1]/div[2]/div[12]/div/div/div/a/div[2]/div[1]/div[2]/ul/li[2]'
        # camera = driver.find_element_by_xpath(
        #     f'//*[@id="container"]/div/div[3]/div[2]/div[1]/div[2]/div[{i}]/div/div/div/a/div[2]/div[1]/div[3]/ul/li[3]')
        # battery = driver.find_element_by_xpath(
        #     f'//*[@id="container"]/div/div[3]/div[2]/div[1]/div[2]/div[{i}]/div/div/div/a/div[2]/div[1]/div[3]/ul/li[4]')
        # price = driver.find_element_by_xpath(
        #     f'//*[@id="container"]/div/div[3]/div[2]/div[1]/div[2]/div[{i}]/div/div/div/a/div[2]/div[2]/div[1]/div/div[1]')
        # sheet.cell(row=i+20*b, column=1).value = mobilename.text
        print(inch.text)
        # print(mobilename.text, ramrom.text, inch.text,
        #       camera.text, battery.text, price.text)
    time.sleep(2)
    driver.find_element_by_link_text(str(a)).click()
    time.sleep(2)
    a += 1
    b += 1
    # workbook.save("C:/Users/HP/Desktop/try.xlsx")
driver.close()
