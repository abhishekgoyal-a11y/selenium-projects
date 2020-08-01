
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
import openpyxl
import os


def line_space_remover(text):
    space = '\n'
    charnumber = 0
    number = 0
    num = 0
    lists = []
    word = " "
    for char in text:
        charnumber += 1
        if char in space:
            lists.append(word[number:charnumber])
            number = charnumber-num
            num += 1
        else:
            word += text[charnumber-1]
    return lists


def youtube_trending_videos():
    a = open("C:/Users/HP/Desktop/youtube trending videos.txt",
             "w", encoding='utf-8')
    path = 'C:/Users/HP/Desktop/youtube trending videos.xlsx'
    details = ['Video name', 'Channel', 'Views', 'Time', 'link']
    w = openpyxl.load_workbook(path)
    sheet = w.active
    rows = sheet.max_row
    cols = sheet.max_column
    driver = webdriver.Chrome(
        'C:/Users/HP/Desktop/Drivers/chromedriver_win32/chromedriver.exe')
    driver.get('https://www.youtube.com/feed/trending')
    name = driver.find_elements_by_class_name(
        'style-scope ytd-video-renderer')
    for i in range(0, 21):
        name_link = driver.find_element_by_xpath(
            f'//*[@id="grid-container"]/ytd-video-renderer[{i+1}]').find_element_by_css_selector('a[id=thumbnail]')
        texts = line_space_remover(name[i].text)
        try:
            del texts[texts.index('•')]
        except:
            pass
        for c in range(len(texts)):
            sheet.cell(row=i+1, column=c+1).value = texts[c]
            a.write(f"{details[c]}:-{texts[c]}")
            a.write('\n')
            # print(texts[c])
        sheet.cell(row=i+1, column=c+2).value = name_link.get_attribute('href')
        a.write(f"{details[4]}:-{name_link.get_attribute('href')}")
        a.write('\n')
        # print(name_link.get_attribute('href'))
    w.save(path)
    driver.close()
